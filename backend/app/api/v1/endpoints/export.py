"""
数据导出路由模块
提供Excel、PDF、JSON等格式的工程量数据导出功能
"""

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List, Dict, Any
import logging
import io
import json
from datetime import datetime

from ....database import get_db
from ....models.drawing import Drawing
from ...deps import get_current_user
from ....models.user import User

logger = logging.getLogger(__name__)

router = APIRouter()

def get_drawing_export_data(drawing_id: int, db: Session, user_id: int) -> Dict[str, Any]:
    """获取图纸导出数据"""
    drawing = db.query(Drawing).filter(
        Drawing.id == drawing_id,
        Drawing.user_id == user_id
    ).first()
    
    if not drawing:
        raise HTTPException(status_code=404, detail="图纸不存在")
    
    if not drawing.processing_result:
        raise HTTPException(status_code=400, detail="图纸未处理完成，无法导出")
    
    try:
        if isinstance(drawing.processing_result, str):
            result = json.loads(drawing.processing_result)
        else:
            result = drawing.processing_result
        
        components = result.get('final_components', [])
        
        # 构建导出数据结构
        export_data = {
            'drawing_info': {
                'id': drawing.id,
                'filename': drawing.filename,
                'status': drawing.status,
                'created_at': drawing.created_at.isoformat() if drawing.created_at else None,
                'updated_at': drawing.updated_at.isoformat() if drawing.updated_at else None,
                'file_size': drawing.file_size,
                'components_count': len(components)
            },
            'components': [],
            'materials': {},
            'summary': {
                'total_components': len(components),
                'total_volume': 0,
                'total_area': 0,
                'by_type': {},
                'by_material': {}
            }
        }
        
        # 处理构件数据
        for i, comp in enumerate(components, 1):
            component_data = {
                '序号': i,
                '构件编号': comp.get('code', ''),
                '构件类型': comp.get('type', ''),
                '尺寸规格': comp.get('dimensions', ''),
                '材料': comp.get('material', ''),
                '数量': comp.get('quantity', 1),
                '单位': comp.get('unit', ''),
                '体积(m³)': comp.get('volume', 0),
                '面积(m²)': comp.get('area', 0),
                '长度(m)': comp.get('length', 0),
                '备注': comp.get('notes', '')
            }
            export_data['components'].append(component_data)
            
            # 统计汇总
            comp_type = comp.get('type', 'unknown')
            material = comp.get('material', 'unknown')
            
            export_data['summary']['by_type'][comp_type] = export_data['summary']['by_type'].get(comp_type, 0) + 1
            export_data['summary']['by_material'][material] = export_data['summary']['by_material'].get(material, 0) + 1
            
            if comp.get('volume'):
                export_data['summary']['total_volume'] += float(comp.get('volume', 0))
            if comp.get('area'):
                export_data['summary']['total_area'] += float(comp.get('area', 0))
            
            # 材料汇总
            if material != 'unknown':
                if material not in export_data['materials']:
                    export_data['materials'][material] = {
                        '材料名称': material,
                        '使用数量': 0,
                        '总体积(m³)': 0,
                        '总面积(m²)': 0,
                        '构件类型': set()
                    }
                
                export_data['materials'][material]['使用数量'] += comp.get('quantity', 1)
                export_data['materials'][material]['总体积(m³)'] += float(comp.get('volume', 0))
                export_data['materials'][material]['总面积(m²)'] += float(comp.get('area', 0))
                export_data['materials'][material]['构件类型'].add(comp_type)
        
        # 转换材料数据格式
        materials_list = []
        for material_name, material_data in export_data['materials'].items():
            materials_list.append({
                '材料名称': material_name,
                '使用数量': material_data['使用数量'],
                '总体积(m³)': round(material_data['总体积(m³)'], 3),
                '总面积(m²)': round(material_data['总面积(m²)'], 3),
                '适用构件': ', '.join(material_data['构件类型'])
            })
        export_data['materials'] = materials_list
        
        # 汇总数据
        summary_list = [
            {'项目': '总构件数', '数值': export_data['summary']['total_components'], '单位': '个'},
            {'项目': '总体积', '数值': round(export_data['summary']['total_volume'], 3), '单位': 'm³'},
            {'项目': '总面积', '数值': round(export_data['summary']['total_area'], 3), '单位': 'm²'},
        ]
        
        for comp_type, count in export_data['summary']['by_type'].items():
            summary_list.append({'项目': f'{comp_type}构件', '数值': count, '单位': '个'})
        
        export_data['summary'] = summary_list
        
        return export_data
        
    except (json.JSONDecodeError, AttributeError) as e:
        logger.error(f"解析图纸 {drawing_id} 数据失败: {e}")
        raise HTTPException(status_code=500, detail="图纸数据格式错误")

@router.get("/excel/{drawing_id}")
async def export_to_excel(
    drawing_id: int,
    include_charts: bool = Query(True, description="是否包含图表"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出为Excel格式"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
        
        # 获取导出数据
        export_data = get_drawing_export_data(drawing_id, db, current_user.id)
        
        # 创建Excel工作簿
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # 构件清单工作表
            if export_data['components']:
                components_df = pd.DataFrame(export_data['components'])
                components_df.to_excel(writer, sheet_name='构件清单', index=False)
            
            # 材料清单工作表
            if export_data['materials']:
                materials_df = pd.DataFrame(export_data['materials'])
                materials_df.to_excel(writer, sheet_name='材料清单', index=False)
            
            # 工程量汇总工作表
            if export_data['summary']:
                summary_df = pd.DataFrame(export_data['summary'])
                summary_df.to_excel(writer, sheet_name='工程量汇总', index=False)
            
            # 图纸信息工作表
            drawing_info = [
                ['图纸ID', export_data['drawing_info']['id']],
                ['文件名', export_data['drawing_info']['filename']],
                ['处理状态', export_data['drawing_info']['status']],
                ['创建时间', export_data['drawing_info']['created_at']],
                ['更新时间', export_data['drawing_info']['updated_at']],
                ['文件大小', export_data['drawing_info']['file_size']],
                ['构件总数', export_data['drawing_info']['components_count']],
                ['导出时间', datetime.now().isoformat()]
            ]
            drawing_df = pd.DataFrame(drawing_info, columns=['项目', '值'])
            drawing_df.to_excel(writer, sheet_name='图纸信息', index=False)
            
            # 美化Excel格式
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # 设置标题行样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        
        # 生成文件名
        filename = f"工程量清单_{export_data['drawing_info']['filename']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少 pandas 或 openpyxl 依赖包")
    except Exception as e:
        logger.error(f"Excel导出失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Excel导出失败: {str(e)}")

@router.get("/json/{drawing_id}")
async def export_to_json(
    drawing_id: int,
    pretty: bool = Query(True, description="是否格式化JSON"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出为JSON格式"""
    try:
        # 获取导出数据
        export_data = get_drawing_export_data(drawing_id, db, current_user.id)
        
        # 添加导出元信息
        export_data['export_meta'] = {
            'export_time': datetime.now().isoformat(),
            'export_user': current_user.email,
            'export_version': '1.0',
            'format': 'json'
        }
        
        # 生成JSON字符串
        if pretty:
            json_str = json.dumps(export_data, ensure_ascii=False, indent=2)
        else:
            json_str = json.dumps(export_data, ensure_ascii=False)
        
        # 生成文件名
        filename = f"工程量数据_{export_data['drawing_info']['filename']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        return Response(
            content=json_str,
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"JSON导出失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"JSON导出失败: {str(e)}")

@router.get("/csv/{drawing_id}")
async def export_to_csv(
    drawing_id: int,
    table: str = Query("components", description="导出表格 (components/materials/summary)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出为CSV格式"""
    try:
        import pandas as pd
        
        # 获取导出数据
        export_data = get_drawing_export_data(drawing_id, db, current_user.id)
        
        # 选择要导出的数据表
        if table == "components" and export_data['components']:
            df = pd.DataFrame(export_data['components'])
            table_name = "构件清单"
        elif table == "materials" and export_data['materials']:
            df = pd.DataFrame(export_data['materials'])
            table_name = "材料清单"
        elif table == "summary" and export_data['summary']:
            df = pd.DataFrame(export_data['summary'])
            table_name = "工程量汇总"
        else:
            raise HTTPException(status_code=400, detail=f"无效的表格类型: {table}")
        
        # 生成CSV
        buffer = io.StringIO()
        df.to_csv(buffer, index=False, encoding='utf-8-sig')  # 使用utf-8-sig以支持中文
        
        # 生成文件名
        filename = f"{table_name}_{export_data['drawing_info']['filename']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return Response(
            content=buffer.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少 pandas 依赖包")
    except Exception as e:
        logger.error(f"CSV导出失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"CSV导出失败: {str(e)}")

@router.get("/pdf/{drawing_id}")
async def export_to_pdf(
    drawing_id: int,
    template: str = Query("standard", description="报告模板 (standard/detailed/summary)"),
    include_charts: bool = Query(True, description="是否包含图表"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出为PDF报告"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        # 获取导出数据
        export_data = get_drawing_export_data(drawing_id, db, current_user.id)
        
        # 创建PDF缓冲区
        buffer = io.BytesIO()
        
        # 创建PDF文档
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # 添加中文字体支持 (如果可用)
        try:
            # 这里需要系统有中文字体文件，实际使用时需要配置正确的字体路径
            pdfmetrics.registerFont(TTFont('SimSun', 'SimSun.ttf'))
            chinese_style = ParagraphStyle(
                'Chinese',
                parent=styles['Normal'],
                fontName='SimSun',
                fontSize=10
            )
        except:
            chinese_style = styles['Normal']
        
        # 构建PDF内容
        story = []
        
        # 标题
        title = Paragraph(f"工程量分析报告 - {export_data['drawing_info']['filename']}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # 图纸信息
        info_data = [
            ['图纸ID', str(export_data['drawing_info']['id'])],
            ['文件名', export_data['drawing_info']['filename']],
            ['处理状态', export_data['drawing_info']['status']],
            ['构件总数', str(export_data['drawing_info']['components_count'])],
            ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        info_table = Table(info_data)
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(Paragraph("图纸基本信息", styles['Heading2']))
        story.append(info_table)
        story.append(Spacer(1, 12))
        
        # 工程量汇总
        if export_data['summary']:
            story.append(Paragraph("工程量汇总", styles['Heading2']))
            summary_data = [['项目', '数值', '单位']]
            summary_data.extend([[item['项目'], str(item['数值']), item['单位']] for item in export_data['summary']])
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 12))
        
        # 根据模板添加详细内容
        if template in ["standard", "detailed"]:
            # 构件清单
            if export_data['components']:
                story.append(Paragraph("构件清单", styles['Heading2']))
                
                # 限制显示的构件数量以避免PDF过大
                max_components = 50 if template == "standard" else len(export_data['components'])
                components_to_show = export_data['components'][:max_components]
                
                comp_data = [['序号', '构件编号', '类型', '尺寸', '材料', '数量', '单位']]
                comp_data.extend([
                    [
                        str(comp['序号']),
                        comp['构件编号'],
                        comp['构件类型'],
                        comp['尺寸规格'],
                        comp['材料'],
                        str(comp['数量']),
                        comp['单位']
                    ] for comp in components_to_show
                ])
                
                comp_table = Table(comp_data)
                comp_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(comp_table)
                
                if len(export_data['components']) > max_components:
                    story.append(Spacer(1, 6))
                    story.append(Paragraph(f"注：仅显示前{max_components}个构件，完整数据请下载Excel版本", chinese_style))
        
        # 构建PDF
        doc.build(story)
        buffer.seek(0)
        
        # 生成文件名
        filename = f"工程量报告_{export_data['drawing_info']['filename']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type='application/pdf',
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少 reportlab 依赖包")
    except Exception as e:
        logger.error(f"PDF导出失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"PDF导出失败: {str(e)}")

@router.get("/templates")
async def get_export_templates():
    """获取导出模板列表"""
    try:
        templates = {
            "excel": {
                "name": "Excel工程量清单",
                "description": "包含构件清单、材料清单、工程量汇总等多个工作表",
                "format": "xlsx",
                "supports_charts": True,
                "file_size": "通常 < 1MB"
            },
            "json": {
                "name": "JSON数据格式",
                "description": "结构化数据，适合程序处理和数据交换",
                "format": "json",
                "supports_charts": False,
                "file_size": "通常 < 500KB"
            },
            "csv": {
                "name": "CSV表格数据",
                "description": "纯表格数据，可选择构件、材料或汇总表",
                "format": "csv",
                "supports_charts": False,
                "file_size": "通常 < 200KB"
            },
            "pdf_standard": {
                "name": "标准PDF报告",
                "description": "包含图纸信息、工程量汇总和前50个构件",
                "format": "pdf",
                "supports_charts": True,
                "file_size": "通常 < 2MB"
            },
            "pdf_detailed": {
                "name": "详细PDF报告",
                "description": "包含完整的构件清单和材料清单",
                "format": "pdf",
                "supports_charts": True,
                "file_size": "通常 < 5MB"
            },
            "pdf_summary": {
                "name": "摘要PDF报告",
                "description": "仅包含图纸信息和工程量汇总",
                "format": "pdf",
                "supports_charts": True,
                "file_size": "通常 < 500KB"
            }
        }
        
        return {
            "success": True,
            "data": templates,
            "message": "成功获取导出模板列表"
        }
        
    except Exception as e:
        logger.error(f"获取导出模板失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取导出模板失败: {str(e)}")

@router.post("/batch/{format}")
async def batch_export(
    format: str,
    drawing_ids: List[int],
    template: Optional[str] = "standard",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """批量导出多个图纸"""
    try:
        if format not in ["excel", "json", "csv"]:
            raise HTTPException(status_code=400, detail="批量导出仅支持 excel, json, csv 格式")
        
        if len(drawing_ids) > 10:
            raise HTTPException(status_code=400, detail="批量导出最多支持10个图纸")
        
        # 收集所有图纸数据
        all_export_data = []
        for drawing_id in drawing_ids:
            try:
                export_data = get_drawing_export_data(drawing_id, db, current_user.id)
                all_export_data.append(export_data)
            except HTTPException as e:
                logger.warning(f"图纸 {drawing_id} 导出数据获取失败: {e.detail}")
                continue
        
        if not all_export_data:
            raise HTTPException(status_code=400, detail="没有可导出的图纸数据")
        
        # 根据格式进行批量导出
        if format == "json":
            # JSON格式：合并所有数据
            batch_data = {
                "export_meta": {
                    "export_time": datetime.now().isoformat(),
                    "export_user": current_user.email,
                    "batch_count": len(all_export_data),
                    "format": "json_batch"
                },
                "drawings": all_export_data
            }
            
            json_str = json.dumps(batch_data, ensure_ascii=False, indent=2)
            filename = f"批量工程量数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            return Response(
                content=json_str,
                media_type="application/json",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        else:
            # Excel和CSV格式
            import pandas as pd
            
            # 合并所有构件数据
            all_components = []
            all_materials = []
            all_summaries = []
            
            for export_data in all_export_data:
                # 添加图纸信息到每行数据
                for comp in export_data['components']:
                    comp['图纸名称'] = export_data['drawing_info']['filename']
                    comp['图纸ID'] = export_data['drawing_info']['id']
                    all_components.append(comp)
                
                for material in export_data['materials']:
                    material['图纸名称'] = export_data['drawing_info']['filename']
                    material['图纸ID'] = export_data['drawing_info']['id']
                    all_materials.append(material)
                
                for summary in export_data['summary']:
                    summary['图纸名称'] = export_data['drawing_info']['filename']
                    summary['图纸ID'] = export_data['drawing_info']['id']
                    all_summaries.append(summary)
            
            if format == "excel":
                # Excel格式
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    if all_components:
                        pd.DataFrame(all_components).to_excel(writer, sheet_name='批量构件清单', index=False)
                    if all_materials:
                        pd.DataFrame(all_materials).to_excel(writer, sheet_name='批量材料清单', index=False)
                    if all_summaries:
                        pd.DataFrame(all_summaries).to_excel(writer, sheet_name='批量工程量汇总', index=False)
                
                buffer.seek(0)
                filename = f"批量工程量清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                return StreamingResponse(
                    io.BytesIO(buffer.read()),
                    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={"Content-Disposition": f"attachment; filename={filename}"}
                )
            
            elif format == "csv":
                # CSV格式：仅导出构件清单
                if not all_components:
                    raise HTTPException(status_code=400, detail="没有构件数据可导出")
                
                df = pd.DataFrame(all_components)
                buffer = io.StringIO()
                df.to_csv(buffer, index=False, encoding='utf-8-sig')
                
                filename = f"批量构件清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                
                return Response(
                    content=buffer.getvalue(),
                    media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"}
                )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少 pandas 依赖包")
    except Exception as e:
        logger.error(f"批量导出失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"批量导出失败: {str(e)}") 