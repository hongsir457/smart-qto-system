# ChatGPT智能工程量分析服务
import os
import base64
import json
import logging
from typing import Dict, List, Any, Optional
from pathlib import Path
import requests
from PIL import Image
import io
import time
from ..config.model_config import ModelConfig

logger = logging.getLogger(__name__)

class ChatGPTQuantityAnalyzer:
    """
    基于ChatGPT-4V的智能施工图工程量分析器
    
    核心功能：
    1. 将PDF图纸转换为高质量图像
    2. 通过ChatGPT Vision API智能识别构件
    3. 按照工程量清单计算规则统计工程量
    4. 输出标准化的工程量清单表格
    """
    
    def __init__(self, api_key: str = None, api_base: str = None):
        """
        初始化ChatGPT分析器
        
        Args:
            api_key: OpenAI API密钥
            api_base: API基础URL（支持代理）
        """
        self.api_key = api_key or os.getenv('OPENAI_API_KEY')
        self.api_base = api_base or os.getenv('OPENAI_API_BASE', 'https://api.openai.com/v1')
        
        if not self.api_key:
            raise ValueError("必须提供OpenAI API密钥")
        
        logger.info(f"当前使用的API Key (末尾4位): ...{self.api_key[-4:]}") # 记录正在使用的API Key最后四位
        logger.info(f"当前使用的API Base: {self.api_base}")

        self.headers = {
            'Authorization': f'Bearer {self.api_key}',
            'Content-Type': 'application/json'
        }
        
        # 优化后的工程量计算提示模板
        self.quantity_prompt_template = """
你是一位资深的建筑工程师和造价师，拥有20年的施工图识图和工程量计算经验。请仔细观察这张建筑图纸，运用你的专业知识进行详细分析。

**重要要求：**
1. 你必须仔细观察图纸中的每一个细节
2. 识别所有可见的文字、数字、标注和符号
3. 不要猜测或编造信息，只基于图纸中实际可见的内容
4. 如果某些信息不清晰，在备注中说明具体原因
5. 必须返回标准JSON格式

**详细分析步骤：**

**第一步：图纸基本信息识别**
- 仔细查看图纸标题栏（通常在右下角）
- 识别项目名称、图纸名称、图号、比例、设计单位等
- 查看图纸中的文字说明和图例

**第二步：构件识别与分类**
- **墙体识别**：粗实线（通常200-300mm宽）表示承重墙，细实线表示非承重墙
- **柱子识别**：方形或圆形截面，通常有尺寸标注（如500×500、Φ400等）
- **梁识别**：连接柱子的线条，查看截面尺寸标注（如300×600等）
- **门窗识别**：门用弧线表示开启方向，窗用细线表示
- **楼板识别**：大面积区域，可能有厚度标注

**第三步：尺寸和数量统计**
- 仔细读取所有尺寸标注（单位通常为mm）
- 统计相同类型构件的数量
- 计算构件的长度、面积或体积
- 注意轴线编号和定位尺寸

**第四步：构件编号识别**
- 查找构件编号（如KZ1、KL1、Q1等）
- 识别构件规格型号
- 注意材料等级标注（如C30混凝土、HRB400钢筋等）

**输出要求：**
请严格按照以下JSON格式输出，不要添加任何解释文字：

```json
{
  "project_info": {
    "project_name": "从图纸标题栏中提取的真实项目名称，如果看不清楚请写'图纸不清晰，无法识别'",
    "drawing_name": "图纸类型（如'一层平面图'、'基础平面图'、'结构平面图'等）",
    "drawing_number": "图号（如'A-01'、'S-02'等），如果看不到请写'无图号标注'",
    "scale": "比例（如'1:100'、'1:200'等），如果看不到请写'无比例标注'",
    "design_stage": "设计阶段（如'施工图'、'方案图'等）"
  },
  "quantity_list": [
    {
      "sequence": 1,
      "component_type": "具体构件类型（如'框架柱'、'剪力墙'、'框架梁'、'现浇板'等）",
      "component_code": "构件编号（如'KZ-1'、'KL-1'等，如果图纸中没有编号，请根据构件类型自行编号）",
      "component_count": "该类型构件的数量（整数）",
      "section_size": "构件截面尺寸（如'500×500'、'300×600'、'厚120'等，必须基于图纸标注）",
      "project_name": "工程量清单项目名称（如'现浇混凝土柱'、'砌体墙'等）",
      "unit": "计量单位（m³、m²、m、个等）",
      "quantity": "工程量数值（小数，基于实际计算）",
      "calculation_formula": "计算公式（如'截面积×高度×数量'、'长度×高度×厚度'等）",
      "remarks": "详细说明（包括具体在图纸哪个位置看到的，尺寸是否清晰，是否有疑问等）",
      "source_page": 1,
      "source_pages": [1]
    }
  ],
  "summary": {
    "total_items": "识别到的构件类型总数",
    "main_structure_volume": "主体结构混凝土体积（m³）",
    "steel_reinforcement_weight": "钢筋重量估算（t）",
    "formwork_area": "模板面积估算（m²）",
    "analysis_confidence": "分析可信度（0.0-1.0，基于图纸清晰度和信息完整度）",
    "missing_information": ["列出图纸中缺失或不清晰的信息"]
  }
}
```

**特别注意：**
- 如果图纸模糊或信息不全，请在remarks中详细说明
- 不要编造尺寸数据，只使用图纸中实际可见的标注
- 如果无法确定某个构件的具体信息，请在备注中说明原因
- 分析可信度要根据图纸质量和信息完整度客观评估

现在请开始分析这张图纸：
"""

    def analyze_drawing_pdf(self, pdf_path: str, project_context: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        分析单张PDF施工图
        
        Args:
            pdf_path: PDF文件路径
            project_context: 项目上下文信息
            
        Returns:
            包含工程量分析结果的字典
        """
        try:
            logger.info(f"开始分析施工图PDF: {pdf_path}")

            # --- 新增：执行文本API预检查 ---
            logger.info("执行文本API预检查...")
            text_api_success = self.test_text_api_call()
            if not text_api_success:
                logger.error("文本API预检查失败。无法继续进行图像分析。")
                return {
                    "error": "基本文本API调用失败，请检查API密钥、权限和网络连接。",
                    "project_info": {},
                    "quantity_list": [],
                    "summary": {"analysis_confidence": 0.0, "failure_stage": "text_api_precheck"}
                }
            logger.info("文本API预检查成功。继续进行图像分析...")
            # --- 预检查结束 ---
            
            # 1. 将PDF转换为高质量图像
            images = self._convert_pdf_to_images(pdf_path)
            if not images:
                raise ValueError("PDF转换为图像失败")
            
            # 2. 分析每页图纸
            all_results = []
            for i, image in enumerate(images):
                logger.info(f"分析第 {i+1} 页图纸")
                
                # 编码图像为base64
                image_base64 = self._encode_image_to_base64(image)
                
                # 构建分析提示
                prompt = self._build_analysis_prompt(project_context)
                
                # 调用ChatGPT API
                result = self._call_chatgpt_vision_api(prompt, image_base64)
                
                if result:
                    result['page_number'] = i + 1
                    all_results.append(result)
            
            # 3. 汇总所有页面的分析结果
            final_result = self._merge_analysis_results(all_results)
            
            logger.info(f"完成PDF分析，共处理 {len(images)} 页，识别到 {len(final_result.get('quantity_list', []))} 个工程量项目")
            
            return final_result
            
        except Exception as e:
            logger.error(f"分析PDF时发生错误: {e}")
            return {
                "error": str(e),
                "project_info": {},
                "quantity_list": [],
                "summary": {"analysis_confidence": 0.0, "failure_stage": "pdf_analysis_exception"}
            }

    def _convert_pdf_to_images(self, pdf_path: str, dpi: int = 300) -> List[Image.Image]:
        """
        将PDF转换为高质量图像
        
        Args:
            pdf_path: PDF文件路径
            dpi: 图像分辨率
            
        Returns:
            图像列表
        """
        try:
            import fitz  # PyMuPDF
            
            images = []
            pdf_document = fitz.open(pdf_path)
            
            for page_num in range(pdf_document.page_count):
                page = pdf_document[page_num]
                
                # 设置缩放矩阵以获得高分辨率
                zoom = dpi / 72  # 72是PDF的默认DPI
                mat = fitz.Matrix(zoom, zoom)
                
                # 渲染页面为图像
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                
                # 转换为PIL图像
                image = Image.open(io.BytesIO(img_data))
                images.append(image)
                
                logger.debug(f"PDF第 {page_num + 1} 页转换完成，图像尺寸: {image.size}")
            
            pdf_document.close()
            return images
            
        except ImportError:
            logger.error("PyMuPDF未安装，请运行: pip install PyMuPDF")
            return []
        except Exception as e:
            logger.error(f"PDF转换失败: {e}")
            return []

    def _encode_image_to_base64(self, image: Image.Image, quality: int = 85) -> str:
        """
        将PIL图像编码为base64字符串
        
        Args:
            image: PIL图像对象
            quality: JPEG压缩质量
            
        Returns:
            base64编码的图像字符串
        """
        try:
            # 如果图像过大，适当压缩
            max_size = (2048, 2048)
            if image.size[0] > max_size[0] or image.size[1] > max_size[1]:
                image = image.copy()
                image.thumbnail(max_size, Image.Resampling.LANCZOS)
                logger.debug(f"图像已压缩至: {image.size}")
            
            # 转换为RGB模式（如果不是）
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # 编码为JPEG并转base64
            buffer = io.BytesIO()
            image.save(buffer, format='JPEG', quality=quality, optimize=True)
            img_str = base64.b64encode(buffer.getvalue()).decode()
            
            return img_str
            
        except Exception as e:
            logger.error(f"图像编码失败: {e}")
            return ""

    def _build_analysis_prompt(self, project_context: Dict[str, Any] = None) -> str:
        """
        构建分析提示词
        
        Args:
            project_context: 项目上下文信息
            
        Returns:
            完整的提示词
        """
        prompt = self.quantity_prompt_template
        
        if project_context:
            context_info = f"""
**项目背景信息：**
- 项目名称: {project_context.get('project_name', '未知')}
- 建筑类型: {project_context.get('building_type', '未知')}
- 结构类型: {project_context.get('structure_type', '框架结构')}
- 设计阶段: {project_context.get('design_stage', '施工图设计')}
- 特殊要求: {project_context.get('special_requirements', '无')}

"""
            prompt = context_info + prompt
        
        return prompt

    def _call_chatgpt_vision_api(self, prompt: str, image_base64: str, model: str = None) -> Optional[Dict[str, Any]]:
        """
        调用ChatGPT Vision API，增加重试机制和更好的错误处理
        """
        max_retries = 3
        base_delay = 2  # 秒

        for attempt in range(max_retries):
            try:
                # 如果没有指定模型，使用配置文件中的默认视觉模型
                if model is None:
                    model = ModelConfig.get_vision_model()
                
                payload = {
                    "model": model,
                    "messages": [
                        {
                            "role": "system",
                            "content": "你是一位专业的建筑工程师和造价师。请仔细分析图纸，提供准确的工程量信息。如果信息不清晰，请在备注中详细说明。"
                        },
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": prompt
                                },
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:image/jpeg;base64,{image_base64}",
                                        "detail": "high"
                                    }
                                }
                            ]
                        }
                    ],
                    "max_tokens": 4000,
                    "temperature": 0.1,  # 降低温度以获得更一致的结果
                    "top_p": 0.9,
                    "frequency_penalty": 0.0,
                    "presence_penalty": 0.0
                }
                
                logger.info(f"尝试调用ChatGPT Vision API (尝试 {attempt + 1}/{max_retries})，模型: {model}")
                response = requests.post(
                    f"{self.api_base}/chat/completions",
                    headers=self.headers,
                    json=payload,
                    timeout=180  # 增加超时到3分钟
                )
                
                # 检查是否为明确的客户端或服务器端错误
                if 400 <= response.status_code < 500:
                    logger.error(f"ChatGPT API客户端错误: {response.status_code}, {response.text}")
                    # 对于404 model_not_found，不再重试
                    if response.status_code == 404 and "model_not_found" in response.text:
                         raise requests.exceptions.RequestException(f"模型 {model} 未找到或已弃用。")
                    # 对于401认证错误，也不重试
                    if response.status_code == 401:
                        raise requests.exceptions.RequestException("API密钥无效或未授权")
                
                response.raise_for_status()  # 对于4xx/5xx错误，这将抛出HTTPError
                
                result = response.json()
                content = result['choices'][0]['message']['content']
                
                logger.info(f"ChatGPT API调用成功，响应长度: {len(content)} 字符")
                
                try:
                    # 尝试多种方式解析JSON
                    json_content = self._extract_json_from_response(content)
                    parsed_result = json.loads(json_content)
                    
                    # 验证和修复解析结果
                    parsed_result = self._validate_and_fix_result(parsed_result)
                    
                    logger.info(f"成功解析ChatGPT响应，识别到 {len(parsed_result.get('quantity_list', []))} 个工程量项目")
                    return parsed_result
                    
                except json.JSONDecodeError as e:
                    logger.warning(f"JSON解析失败: {e}")
                    logger.debug(f"原始响应内容: {content[:1000]}...")
                    
                    # 如果是最后一次尝试，创建回退结果
                    if attempt == max_retries - 1:
                        logger.error("所有尝试均失败，创建回退结果")
                        fallback_result = self._create_fallback_result(content)
                        return fallback_result
                    else:
                        logger.info(f"JSON解析失败，将重试 (尝试 {attempt + 2}/{max_retries})")
                        continue
            
            except requests.exceptions.Timeout:
                logger.warning(f"ChatGPT API调用超时 (尝试 {attempt + 1}/{max_retries})")
                if attempt == max_retries - 1:
                    logger.error("ChatGPT API调用最终超时")
                    raise  # 重新抛出最后的超时异常
                time.sleep(base_delay * (2 ** attempt)) # 指数退避
            except requests.exceptions.RequestException as e:
                # 包括 ConnectionError, HTTPError, TooManyRedirects 等
                logger.warning(f"ChatGPT API请求失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt == max_retries - 1:
                    logger.error(f"ChatGPT API请求最终失败: {e}")
                    raise # 重新抛出最后的请求异常
                time.sleep(base_delay * (2 ** attempt)) # 指数退避
            except Exception as e:
                logger.error(f"调用ChatGPT API时发生其他错误 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt == max_retries - 1:
                    logger.error(f"调用ChatGPT API时最终发生其他错误: {e}")
                    raise # 重新抛出最后的未知异常
                time.sleep(base_delay * (2 ** attempt)) # 指数退避
        
        logger.error("所有重试尝试后，ChatGPT API调用失败")
        return None

    def _merge_analysis_results(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        合并多页分析结果
        
        Args:
            results: 多页分析结果列表
            
        Returns:
            合并后的最终结果
        """
        if not results:
            return {
                "project_info": {},
                "quantity_list": [],
                "summary": {"analysis_confidence": 0.0}
            }
        
        # 获取项目信息（以第一页为准）
        project_info = results[0].get('project_info', {})
        
        # 合并所有工程量项目
        all_quantity_items = []
        for result in results:
            if 'quantity_list' in result:
                for item in result['quantity_list']:
                    # 添加页码信息
                    item['source_page'] = result.get('page_number', 1)
                    all_quantity_items.append(item)
        
        # 去重和合并相同构件
        merged_items = self._deduplicate_quantity_items(all_quantity_items)
        
        # 计算汇总信息
        summary = self._calculate_summary(merged_items, results)
        
        return {
            "project_info": project_info,
            "quantity_list": merged_items,
            "summary": summary,
            "total_pages": len(results)
        }

    def _deduplicate_quantity_items(self, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        去重和合并相同的工程量项目
        
        Args:
            items: 工程量项目列表
            
        Returns:
            去重后的项目列表
        """
        merged_items = {}
        
        for item in items:
            # 构建唯一键
            key = f"{item.get('component_type', '')}_{item.get('component_code', '')}_{item.get('section_size', '')}"
            
            if key in merged_items:
                # 合并数量
                existing_item = merged_items[key]
                existing_quantity = float(existing_item.get('quantity', 0))
                new_quantity = float(item.get('quantity', 0))
                existing_item['quantity'] = existing_quantity + new_quantity
                
                # 合并源页面信息
                existing_pages = existing_item.get('source_pages', [])
                new_page = item.get('source_page', 0)
                if new_page not in existing_pages:
                    existing_pages.append(new_page)
                existing_item['source_pages'] = existing_pages
            else:
                # 新项目
                item['source_pages'] = [item.get('source_page', 0)]
                merged_items[key] = item
        
        # 重新分配序号
        final_items = list(merged_items.values())
        for i, item in enumerate(final_items, 1):
            item['sequence'] = i
        
        return final_items

    def _calculate_summary(self, items: List[Dict[str, Any]], results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        计算汇总统计信息
        
        Args:
            items: 工程量项目列表
            results: 原始分析结果列表
            
        Returns:
            汇总统计信息
        """
        summary = {
            "total_items": len(items),
            "main_structure_volume": 0.0,
            "steel_reinforcement_weight": 0.0,
            "formwork_area": 0.0,
            "analysis_confidence": 0.0,
            "missing_information": []
        }
        
        # 统计主要工程量
        for item in items:
            unit = item.get('unit', '').lower()
            quantity = float(item.get('quantity', 0))
            component_type = item.get('component_type', '').lower()
            
            # 混凝土工程量（m³）
            if 'm³' in unit or 'm3' in unit:
                if any(keyword in component_type for keyword in ['柱', '梁', '板', '墙', '基础']):
                    summary['main_structure_volume'] += quantity
            
            # 钢筋工程量（kg或t）
            elif ('kg' in unit or 't' in unit) and '钢筋' in component_type:
                if 't' in unit:
                    summary['steel_reinforcement_weight'] += quantity * 1000
                else:
                    summary['steel_reinforcement_weight'] += quantity
            
            # 模板工程量（m²）
            elif 'm²' in unit or 'm2' in unit:
                if '模板' in component_type:
                    summary['formwork_area'] += quantity
        
        # 计算平均置信度
        confidences = [result.get('summary', {}).get('analysis_confidence', 0.0) for result in results]
        if confidences:
            summary['analysis_confidence'] = sum(confidences) / len(confidences)
        
        # 收集缺失信息
        all_missing = []
        for result in results:
            missing = result.get('summary', {}).get('missing_information', [])
            all_missing.extend(missing)
        summary['missing_information'] = list(set(all_missing))
        
        return summary

    def export_to_excel(self, analysis_result: Dict[str, Any], output_path: str) -> bool:
        """
        将分析结果导出为Excel工程量清单
        
        Args:
            analysis_result: 分析结果
            output_path: 输出文件路径
            
        Returns:
            是否导出成功
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "工程量清单"
            
            project_info = analysis_result.get('project_info', {})
            quantity_list = analysis_result.get('quantity_list', [])
            summary = analysis_result.get('summary', {})
            
            # 设置样式
            header_font = Font(bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入项目信息
            row = 1
            ws.merge_cells(f'A{row}:K{row}')
            ws[f'A{row}'] = f"工程量清单 - {project_info.get('project_name', '未知项目')}"
            ws[f'A{row}'].font = Font(bold=True, size=16)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            row += 2
            info_items = [
                ('项目名称', project_info.get('project_name', '')),
                ('设计单位', project_info.get('design_unit', '')),
                ('图纸名称', project_info.get('drawing_name', '')),
                ('图号', project_info.get('drawing_number', '')),
                ('比例', project_info.get('scale', '')),
                ('设计阶段', project_info.get('design_stage', ''))
            ]
            
            for label, value in info_items:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
            
            # 写入表头
            headers = [
                '序号', '图号', '构件类型', '构件编号', '构件数量', 
                '截面尺寸', '工程内容', '计量单位', '统计数量', '计算依据', '备注'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.border = border
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
            
            # 写入数据
            for item in quantity_list:
                data = [
                    item.get('sequence', ''),
                    item.get('drawing_number', ''),
                    item.get('component_type', ''),
                    item.get('component_code', ''),
                    item.get('component_count', ''),
                    item.get('section_size', ''),
                    item.get('project_name', ''),
                    item.get('unit', ''),
                    item.get('quantity', ''),
                    item.get('calculation_formula', ''),
                    item.get('remarks', '')
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col in [1, 5, 9]:  # 数值列右对齐
                        cell.alignment = Alignment(horizontal='right')
                
                row += 1
            
            # 写入汇总信息
            row += 2
            ws[f'A{row}'] = "汇总统计"
            ws[f'A{row}'].font = header_font
            row += 1
            
            summary_items = [
                ('统计项目总数', summary.get('total_items', 0)),
                ('主体结构混凝土总量(m³)', f"{summary.get('main_structure_volume', 0):.2f}"),
                ('钢筋总重量(kg)', f"{summary.get('steel_reinforcement_weight', 0):.2f}"),
                ('模板总面积(m²)', f"{summary.get('formwork_area', 0):.2f}"),
                ('分析置信度', f"{summary.get('analysis_confidence', 0):.1%}")
            ]
            
            for label, value in summary_items:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
            
            # 调整列宽
            column_widths = [6, 12, 15, 12, 8, 15, 20, 8, 12, 20, 25]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width
            
            # 保存文件
            wb.save(output_path)
            logger.info(f"工程量清单已导出到: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            return False

    # 新增的测试方法
    def test_text_api_call(self, model: str = None) -> bool:
        """
        测试一个简单的纯文本API调用
        Args:
            model: 要测试的OpenAI模型
        Returns:
            True 如果API调用成功, False 如果失败
        """
        logger.info(f"开始测试纯文本API调用，目标模型: {model}...")
        
        # 如果没有指定模型，使用配置文件中的默认模型
        if model is None:
            model = ModelConfig.get_default_model()
            
        payload = {
            "model": model,
            "messages": [
                {
                    "role": "user",
                    "content": "Hello, this is a test message. Who are you?"
                }
            ],
            "max_tokens": 100,
            "temperature": 0.7
        }
        
        max_retries = 2 # 对于简单测试，减少重试次数
        base_delay = 1

        for attempt in range(max_retries):
            try:
                logger.info(f"尝试调用文本API (尝试 {attempt + 1}/{max_retries})...")
                response = requests.post(
                    f"{self.api_base}/chat/completions",
                    headers=self.headers,
                    json=payload,
                    timeout=60  # 文本调用超时时间可以短一些
                )
                
                if 400 <= response.status_code < 500:
                    logger.error(f"文本API客户端错误: {response.status_code}, {response.text}")
                    if response.status_code == 401:
                        logger.error("API密钥无效或未授权。请检查您的OPENAI_API_KEY。")
                    elif response.status_code == 404:
                        logger.error(f"模型 {model} 未找到。请确认模型名称是否正确。")
                    return False # 客户端错误，通常不重试
                
                response.raise_for_status()  # 对于其他4xx/5xx错误，这将抛出HTTPError
                
                result = response.json()
                logger.info(f"纯文本API调用成功! 响应: {result.get('choices', [{}])[0].get('message', {}).get('content', 'No content')[:100]}...")
                return True
            
            except requests.exceptions.Timeout:
                logger.warning(f"文本API调用超时 (尝试 {attempt + 1}/{max_retries})")
                if attempt == max_retries - 1:
                    logger.error("文本API调用最终超时")
                    return False
                time.sleep(base_delay * (2 ** attempt))
            except requests.exceptions.RequestException as e:
                logger.warning(f"文本API请求失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt == max_retries - 1:
                    logger.error(f"文本API请求最终失败: {e}")
                    return False
                time.sleep(base_delay * (2 ** attempt))
            except Exception as e:
                logger.error(f"调用文本API时发生其他错误 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt == max_retries - 1:
                    logger.error(f"调用文本API时最终发生其他错误: {e}")
                    return False
                time.sleep(base_delay * (2 ** attempt))
        
        logger.error("所有重试尝试后，文本API调用失败")
        return False

    def _extract_json_from_response(self, content: str) -> str:
        """从响应内容中提取JSON字符串"""
        # 方法1: 查找```json代码块
        if '```json' in content:
            json_start = content.find('```json') + 7
            json_end = content.find('```', json_start)
            if json_end > json_start:
                return content[json_start:json_end].strip()
        
        # 方法2: 查找```代码块（无json标识）
        if '```' in content:
            first_block_start = content.find('```') + 3
            first_block_end = content.find('```', first_block_start)
            if first_block_end > first_block_start:
                potential_json = content[first_block_start:first_block_end].strip()
                if potential_json.startswith('{') and potential_json.endswith('}'):
                    return potential_json
        
        # 方法3: 查找第一个完整的JSON对象
        start_idx = content.find('{')
        if start_idx != -1:
            brace_count = 0
            for i, char in enumerate(content[start_idx:], start_idx):
                if char == '{':
                    brace_count += 1
                elif char == '}':
                    brace_count -= 1
                    if brace_count == 0:
                        return content[start_idx:i+1]
        
        # 方法4: 返回原始内容
        return content

    def _validate_and_fix_result(self, parsed_result: Dict[str, Any]) -> Dict[str, Any]:
        """验证和修复解析结果，提供更严格的质量控制"""
        # 确保必要的字段存在
        if 'project_info' not in parsed_result:
            parsed_result['project_info'] = {}
        
        if 'quantity_list' not in parsed_result:
            parsed_result['quantity_list'] = []
        
        if 'summary' not in parsed_result:
            parsed_result['summary'] = {
                "total_items": len(parsed_result.get('quantity_list', [])),
                "main_structure_volume": 0,
                "steel_reinforcement_weight": 0,
                "formwork_area": 0,
                "analysis_confidence": 0.3,  # 降低默认可信度
                "missing_information": ["图纸信息不完整"]
            }
        
        # 修复项目信息字段 - 更严格的验证
        project_info = parsed_result['project_info']
        default_project_fields = {
            'project_name': '图纸不清晰，无法识别',
            'drawing_name': '图纸类型不明',
            'drawing_number': '无图号标注',
            'scale': '无比例标注',
            'design_stage': '施工图'
        }
        
        for field, default_value in default_project_fields.items():
            if field not in project_info or not project_info[field] or project_info[field] in ['未识别', '']:
                project_info[field] = default_value
        
        # 修复工程量列表 - 更严格的验证
        quantity_list = parsed_result['quantity_list']
        valid_items = []
        
        for i, item in enumerate(quantity_list):
            # 检查必要字段的有效性
            if not item.get('component_type') or item.get('component_type') in ['未知构件', '']:
                logger.warning(f"工程量项目 {i+1} 缺少有效的构件类型，跳过")
                continue
                
            # 确保必要字段存在并有合理值
            if 'sequence' not in item:
                item['sequence'] = len(valid_items) + 1
            if 'component_code' not in item or not item['component_code']:
                item['component_code'] = f"{item.get('component_type', 'UNKNOWN')}_{len(valid_items)+1}"
            if 'component_count' not in item or not isinstance(item.get('component_count'), (int, float)) or item['component_count'] <= 0:
                item['component_count'] = 1
            if 'quantity' not in item or not isinstance(item.get('quantity'), (int, float)) or item['quantity'] <= 0:
                item['quantity'] = 1.0
                item['remarks'] = item.get('remarks', '') + ' [数量信息不明确，设为默认值]'
            if 'unit' not in item or not item['unit']:
                # 根据构件类型推断单位
                component_type = item.get('component_type', '').lower()
                if any(keyword in component_type for keyword in ['墙', '板', '面']):
                    item['unit'] = 'm²'
                elif any(keyword in component_type for keyword in ['柱', '梁', '基础']):
                    item['unit'] = 'm³'
                elif any(keyword in component_type for keyword in ['门', '窗']):
                    item['unit'] = '个'
                else:
                    item['unit'] = 'm'
            if 'project_name' not in item or not item['project_name']:
                item['project_name'] = f"{item.get('component_type', '未知')}工程"
            if 'calculation_formula' not in item or not item['calculation_formula']:
                item['calculation_formula'] = '基于图纸识别计算'
            if 'remarks' not in item:
                item['remarks'] = '需要进一步核实图纸信息'
            elif not item['remarks'] or '推测' in item['remarks']:
                item['remarks'] += ' [建议人工复核]'
                
            # 验证截面尺寸格式
            if 'section_size' not in item or not item['section_size'] or item['section_size'] in ['待确定', '']:
                item['section_size'] = '尺寸待确定'
                item['remarks'] += ' [缺少尺寸标注]'
            
            valid_items.append(item)
        
        parsed_result['quantity_list'] = valid_items
        
        # 更新汇总信息 - 更严格的可信度评估
        summary = parsed_result['summary']
        summary['total_items'] = len(valid_items)
        
        # 计算可信度
        confidence_score = 0.5  # 基础分数
        
        # 项目信息完整性检查
        project_info_score = 0
        for field, value in project_info.items():
            if value and value not in ['图纸不清晰，无法识别', '无图号标注', '无比例标注', '图纸类型不明']:
                project_info_score += 0.1
        confidence_score += min(project_info_score, 0.2)
        
        # 工程量信息质量检查
        if valid_items:
            quality_score = 0
            for item in valid_items:
                item_score = 0
                # 检查是否有具体的尺寸信息
                if item.get('section_size') and '×' in item['section_size'] and item['section_size'] != '尺寸待确定':
                    item_score += 0.3
                # 检查是否有合理的数量
                if isinstance(item.get('quantity'), (int, float)) and item['quantity'] > 0:
                    item_score += 0.2
                # 检查备注是否详细
                if item.get('remarks') and len(item['remarks']) > 10 and '推测' not in item['remarks']:
                    item_score += 0.2
                quality_score += item_score
            
            confidence_score += min(quality_score / len(valid_items), 0.3)
        
        summary['analysis_confidence'] = round(min(confidence_score, 1.0), 2)
        
        # 更新缺失信息列表
        missing_info = []
        if project_info.get('project_name') in ['图纸不清晰，无法识别']:
            missing_info.append('项目名称不清晰')
        if project_info.get('drawing_number') in ['无图号标注']:
            missing_info.append('缺少图号标注')
        if project_info.get('scale') in ['无比例标注']:
            missing_info.append('缺少比例标注')
        
        for item in valid_items:
            if item.get('section_size') == '尺寸待确定':
                missing_info.append(f"{item.get('component_type', '构件')}尺寸不明确")
            if '推测' in item.get('remarks', ''):
                missing_info.append(f"{item.get('component_type', '构件')}信息需要确认")
        
        if not missing_info:
            missing_info = ['建议人工复核图纸']
        
        summary['missing_information'] = list(set(missing_info))  # 去重
        
        return parsed_result

    def _create_fallback_result(self, content: str) -> Dict[str, Any]:
        """从文本内容创建回退结果，提供更好的错误处理"""
        logger.warning("JSON解析失败，尝试从文本中提取信息")
        
        # 尝试从文本中提取一些基本信息
        quantity_list = []
        
        # 更全面的关键词匹配
        construction_keywords = {
            '框架柱': '框架柱',
            '剪力墙': '剪力墙', 
            '框架梁': '框架梁',
            '连续梁': '连续梁',
            '现浇板': '现浇板',
            '预制板': '预制板',
            '基础': '基础',
            '承台': '承台',
            '墙体': '墙体',
            '砌体墙': '砌体墙',
            '门': '门',
            '窗': '窗',
            '柱': '柱',
            '梁': '梁',
            '板': '板',
            '墙': '墙'
        }
        
        lines = content.split('\n')
        sequence = 1
        
        # 尝试提取项目信息
        project_name = '解析失败，无法识别'
        drawing_name = '图纸类型不明'
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # 尝试识别项目名称
            if any(keyword in line for keyword in ['项目', '工程', '建筑']):
                if len(line) < 50:  # 避免提取过长的文本
                    project_name = line
            
            # 尝试识别图纸类型
            if any(keyword in line for keyword in ['平面图', '立面图', '剖面图', '详图']):
                drawing_name = line
                
            # 检查是否包含构件关键词
            for keyword, component_type in construction_keywords.items():
                if keyword in line:
                    # 尝试提取数字（可能是尺寸或数量）
                    import re
                    numbers = re.findall(r'\d+(?:\.\d+)?', line)
                    
                    # 尝试提取尺寸信息
                    size_pattern = r'(\d+)×(\d+)'
                    size_match = re.search(size_pattern, line)
                    section_size = f"{size_match.group(1)}×{size_match.group(2)}" if size_match else "尺寸待确定"
                    
                    quantity_list.append({
                        "sequence": sequence,
                        "component_type": component_type,
                        "component_code": f"{keyword.replace('框架', '').replace('现浇', '')}{sequence}",
                        "component_count": 1,
                        "section_size": section_size,
                        "project_name": f"{component_type}工程",
                        "unit": "m³" if component_type in ['框架柱', '框架梁', '基础'] else "m²",
                        "quantity": float(numbers[0]) if numbers else 1.0,
                        "calculation_formula": "基于文本识别估算",
                        "remarks": f"从响应文本中提取: {line[:50]}... [需要人工核实]",
                        "source_page": 1,
                        "source_pages": [1]
                    })
                    sequence += 1
                    break
        
        return {
            "raw_response": content[:500] + "..." if len(content) > 500 else content,
            "parse_error": "ChatGPT响应无法解析为标准JSON格式",
            "project_info": {
                "project_name": project_name,
                "drawing_name": drawing_name, 
                "drawing_number": "无图号标注",
                "scale": "无比例标注",
                "design_stage": "施工图"
            },
            "quantity_list": quantity_list,
            "summary": {
                "total_items": len(quantity_list),
                "main_structure_volume": 0,
                "steel_reinforcement_weight": 0,
                "formwork_area": 0,
                "analysis_confidence": 0.2,  # 很低的可信度
                "missing_information": [
                    "ChatGPT响应格式错误",
                    "图纸信息提取不完整", 
                    "所有数据需要人工核实"
                ]
            }
        } 