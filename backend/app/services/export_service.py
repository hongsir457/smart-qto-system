#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工程量导出服务
支持Excel格式导出
"""

import logging
import os
import tempfile
from typing import Dict, Any, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

logger = logging.getLogger(__name__)

class ExportService:
    """工程量导出服务"""
    
    def __init__(self):
        """初始化导出服务"""
        if not openpyxl:
            logger.warning("⚠️ openpyxl 库未安装，Excel导出功能不可用")
    
    def export_quantities_to_excel(self, drawing_data: Dict[str, Any]) -> Optional[str]:
        """
        导出工程量到Excel文件
        
        Args:
            drawing_data: 图纸数据，包含工程量计算结果
            
        Returns:
            Optional[str]: Excel文件路径，如果导出失败返回None
        """
        try:
            if not openpyxl:
                raise ImportError("openpyxl 库未安装，无法导出Excel")
            
            logger.info(f"📊 开始导出工程量到Excel: 图纸ID={drawing_data.get('id')}")
            
            # 获取工程量数据
            quantity_results = drawing_data.get('quantity_results', {})
            quantities = quantity_results.get('quantities', {})
            total_summary = quantity_results.get('total_summary', {})
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 删除默认工作表
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # 创建工作表
            self._create_summary_sheet(wb, drawing_data, total_summary)
            
            # 为每种构件类型创建工作表
            component_types = {
                'wall': '墙体工程量',
                'column': '柱子工程量', 
                'beam': '梁工程量',
                'slab': '板工程量',
                'foundation': '基础工程量'
            }
            
            for component_type, sheet_name in component_types.items():
                if component_type in quantities:
                    self._create_component_sheet(
                        wb, sheet_name, component_type, 
                        quantities[component_type]
                    )
            
            # 保存Excel文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"工程量计算_{drawing_data.get('filename', 'unknown')}_{timestamp}.xlsx"
            
            # 创建临时文件
            temp_dir = tempfile.gettempdir()
            file_path = os.path.join(temp_dir, filename)
            
            wb.save(file_path)
            
            logger.info(f"✅ Excel导出完成: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"❌ Excel导出失败: {str(e)}", exc_info=True)
            return None
    
    def _create_summary_sheet(self, wb: Workbook, drawing_data: Dict, total_summary: Dict):
        """创建总工程量汇总表"""
        try:
            ws = wb.create_sheet("总工程量汇总")
            
            # 设置标题样式
            title_font = Font(name="微软雅黑", size=16, bold=True)
            header_font = Font(name="微软雅黑", size=12, bold=True)
            cell_font = Font(name="微软雅黑", size=10)
            
            # 标题
            ws['A1'] = "智能工程量计算系统 - 总工程量汇总表"
            ws['A1'].font = title_font
            ws.merge_cells('A1:F1')
            
            # 项目信息
            row = 3
            info_data = [
                ["图纸名称", drawing_data.get('filename', '')],
                ["计算时间", drawing_data.get('quantity_results', {}).get('calculation_time', '')],
                ["构件总数", total_summary.get('total_count', 0)],
                ["总体积(m³)", f"{total_summary.get('total_volume', 0):.2f}"],
                ["总面积(m²)", f"{total_summary.get('total_area', 0):.2f}"]
            ]
            
            for info in info_data:
                ws[f'A{row}'] = info[0]
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'] = info[1]
                ws[f'B{row}'].font = cell_font
                row += 1
            
            # 分类汇总表头
            row += 2
            headers = ["构件类型", "数量(个)", "体积(m³)", "面积(m²)", "备注"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            # 分类汇总数据
            quantities = drawing_data.get('quantity_results', {}).get('quantities', {})
            component_names = {
                'wall': '墙体',
                'column': '柱子',
                'beam': '梁',
                'slab': '板',
                'foundation': '基础'
            }
            
            row += 1
            for component_type, summary in quantities.items():
                type_summary = summary.get('summary', {})
                
                ws.cell(row=row, column=1, value=component_names.get(component_type, component_type))
                ws.cell(row=row, column=2, value=type_summary.get('count', 0))
                ws.cell(row=row, column=3, value=f"{type_summary.get('volume', 0):.2f}")
                ws.cell(row=row, column=4, value=f"{type_summary.get('area', 0):.2f}")
                ws.cell(row=row, column=5, value="")
                
                row += 1
            
            # 设置列宽
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 20
            
            logger.info("✅ 总工程量汇总表创建完成")
            
        except Exception as e:
            logger.error(f"❌ 创建汇总表失败: {str(e)}")
    
    def _create_component_sheet(self, wb: Workbook, sheet_name: str, component_type: str, data: Dict):
        """创建构件详细工程量表"""
        try:
            ws = wb.create_sheet(sheet_name)
            
            # 设置样式
            title_font = Font(name="微软雅黑", size=14, bold=True)
            header_font = Font(name="微软雅黑", size=10, bold=True)
            cell_font = Font(name="微软雅黑", size=9)
            
            # 标题
            ws['A1'] = f"{sheet_name}明细表"
            ws['A1'].font = title_font
            ws.merge_cells('A1:H1')
            
            # 表头
            headers = [
                "序号", "构件名称", "长度(m)", "宽度(m)", 
                "高度/厚度(m)", "数量", "体积(m³)", "面积(m²)"
            ]
            
            row = 3
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 数据行
            items = data.get('items', [])
            row += 1
            
            for i, item in enumerate(items, 1):
                dimensions = item.get('dimensions', {})
                
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=item.get('name', ''))
                ws.cell(row=row, column=3, value=f"{dimensions.get('length', 0):.2f}")
                ws.cell(row=row, column=4, value=f"{dimensions.get('width', 0):.2f}")
                
                # 根据构件类型显示不同的第三个尺寸
                third_dim = dimensions.get('height', dimensions.get('thickness', 0))
                ws.cell(row=row, column=5, value=f"{third_dim:.2f}")
                
                ws.cell(row=row, column=6, value=item.get('quantity', 0))
                ws.cell(row=row, column=7, value=f"{item.get('volume', 0):.2f}")
                ws.cell(row=row, column=8, value=f"{item.get('area', 0):.2f}")
                
                row += 1
            
            # 汇总行
            summary = data.get('summary', {})
            if summary:
                row += 1
                ws.cell(row=row, column=1, value="汇总")
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=6, value=summary.get('count', 0))
                ws.cell(row=row, column=6).font = header_font
                ws.cell(row=row, column=7, value=f"{summary.get('volume', 0):.2f}")
                ws.cell(row=row, column=7).font = header_font
                ws.cell(row=row, column=8, value=f"{summary.get('area', 0):.2f}")
                ws.cell(row=row, column=8).font = header_font
            
            # 设置列宽
            column_widths = [8, 15, 10, 10, 12, 8, 12, 12]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            logger.info(f"✅ {sheet_name}表创建完成")
            
        except Exception as e:
            logger.error(f"❌ 创建{sheet_name}表失败: {str(e)}")

# 创建全局导出服务实例
export_service = ExportService()